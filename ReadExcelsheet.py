import openpyxl as openpyxl
class ExcelRead:
    Dict = {}
    def fileread(self,sheetName):
        book = openpyxl.load_workbook("C:\\Users\\SaranyaV\\OneDrive - Specialist Lending\\Desktop\\selenium with python\\Countyname.xlsx")
        sheet = book[sheetName]
        print(sheetName)
        cell = sheet.cell(row=1, column=2) # move in to the specific cell
        #print(cell)
        # sheet.cell(row=5, column=2).value = "newvalue"  # write the data in to the excel sheet
        # for mar row
        maxrow = sheet.max_row
        print(maxrow)
        # for max column
        maxcolumnv = sheet.max_column
        print(maxcolumnv)
        for i in range(2, maxrow + 1): # row to column
            for j in range(1, maxcolumnv+1):
        #for i in range(1, maxcolumnv + 1): #column to row
            #for j in range(2, maxrow + 1):
                cell = sheet.cell(row=j, column=i)
                #print(cell.value)
                self.Dict[(sheet.cell(row=1, column=j).value) + str(i-1)] = str(sheet.cell(row=i, column=j).value)

        print(self.Dict)
        return self.Dict

    def filereadmaxrow(self,sheetName):
        book = openpyxl.load_workbook("C:\\Users\\sathishkumar\\PycharmProjects\\WeekdayPython\\inputFile\\InputFile.xlsx")
        sheet = book[sheetName]
        cell = sheet.cell(row=1, column=2) # move in to the specific cell
        print(cell)
        # sheet.cell(row=5, column=2).value = "newvalue"  # write the data in to the excel sheet
        # for mar row
        max_row = sheet.max_row
        return max_row


obj= ExcelRead()
obj.fileread("county")